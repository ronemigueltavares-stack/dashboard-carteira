"""
sync_firebase.py
Lê a planilha "Carteira Investimentos.xlsx" e envia os dados para o Firestore.
Incremental: usa hash MD5 — só grava o que é novo, nunca duplica.

Como rodar:
  python sync_firebase.py
"""

import hashlib
import json
import re
import sys
from datetime import datetime

import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

# ── Configuração ──────────────────────────────────────────────────────────────
EXCEL_FILE = "Carteira Investimentos.xlsx"
KEY_FILE   = "firebase-key.json"
USER_UID   = "3zESUj8nuPN1ViGpyCxnkJGIJYL2"
# ─────────────────────────────────────────────────────────────────────────────

MESES_PT = {
    'jan':1,'fev':2,'mar':3,'abr':4,'mai':5,'jun':6,
    'jul':7,'ago':8,'set':9,'out':10,'nov':11,'dez':12
}

def parse_data(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip().lower()
    # formato "dd-mmm-aa" ex: 06-jun-19
    m = re.match(r'(\d{1,2})-([a-z]{3})-(\d{2,4})$', s)
    if m:
        dia, mes_str, ano = int(m.group(1)), m.group(2), int(m.group(3))
        mes = MESES_PT.get(mes_str)
        if mes:
            ano = ano + 2000 if ano < 100 else ano
            return datetime(ano, mes, dia)
    # formato "mmm-aa" ex: jan-26
    m = re.match(r'([a-z]{3})-(\d{2,4})$', s)
    if m:
        mes_str, ano = m.group(1), int(m.group(2))
        mes = MESES_PT.get(mes_str)
        if mes:
            ano = ano + 2000 if ano < 100 else ano
            return datetime(ano, mes, 1)
    return None


def limpar_valor(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[R$\s]', '', str(v))
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def limpar_pct(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[%\s]', '', str(v)).replace(',', '.')
    try:
        return float(s) / 100
    except ValueError:
        return None


def md5(d):
    return hashlib.md5(
        json.dumps(d, sort_keys=True, default=str).encode()
    ).hexdigest()


def sync():
    print("Conectando ao Firebase...")
    cred = credentials.Certificate(KEY_FILE)
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    lanc_col = db.collection(f'users/{USER_UID}/lancamentos')
    prov_col = db.collection(f'users/{USER_UID}/proventos')

    print("Buscando registros ja existentes no Firestore...")
    hashes_lanc = {d.to_dict().get('hash') for d in lanc_col.stream()}
    hashes_prov = {d.to_dict().get('hash') for d in prov_col.stream()}

    print(f"Lendo planilha: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Ler mapa de classes da aba Cadastro
    def simplificar_tipo(t):
        if not t: return 'Acao'
        t = str(t).strip()
        if t == 'FII':  return 'FII'
        if t == 'ETF':  return 'ETF'
        if t == 'BDR':  return 'BDR'
        if t in ('Fundos', 'Tesouro Direto', 'Renda Fixa'): return 'RF'
        return 'Acao'

    classes = {}
    for row in wb['Cadastro'].iter_rows(min_row=4, values_only=True):
        ticker = row[2]
        tipo   = row[0]
        if ticker and tipo:
            classes[str(ticker).strip().upper()] = simplificar_tipo(tipo)

    ws = wb['Lançamentos']

    lanc_novos = lanc_skip = prov_novos = prov_skip = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        # ── Lançamento (cols A–F) ─────────────────────────────────────────
        data_l  = row[0] if len(row) > 0 else None
        ativo_l = row[1] if len(row) > 1 else None
        qtd_l   = row[2] if len(row) > 2 else None
        valor_l = row[3] if len(row) > 3 else None
        total_l = row[4] if len(row) > 4 else None
        tipo_l  = row[5] if len(row) > 5 else None

        dt_l = parse_data(data_l)
        if ativo_l and tipo_l and dt_l:
            rec = {
                'data':  dt_l,
                'ativo': str(ativo_l).strip().upper(),
                'qtd':   float(qtd_l) if qtd_l is not None else 0.0,
                'valor': limpar_valor(valor_l),
                'total': limpar_valor(total_l),
                'tipo':  str(tipo_l).strip().upper(),
            }
            h = md5({k: str(v) for k, v in rec.items()})
            rec['hash'] = h
            if h not in hashes_lanc:
                lanc_col.document(h).set(rec)
                hashes_lanc.add(h)
                lanc_novos += 1
            else:
                lanc_skip += 1

        # ── Provento (cols H–K) ───────────────────────────────────────────
        data_p  = row[7]  if len(row) > 7  else None
        ativo_p = row[8]  if len(row) > 8  else None
        valor_p = row[9]  if len(row) > 9  else None
        tipo_p  = row[10] if len(row) > 10 else None

        dt_p = parse_data(data_p)
        if ativo_p and tipo_p and dt_p:
            rec = {
                'data':  dt_p,
                'ativo': str(ativo_p).strip().upper(),
                'valor': limpar_valor(valor_p),
                'tipo':  str(tipo_p).strip().upper(),
            }
            h = md5({k: str(v) for k, v in rec.items()})
            rec['hash'] = h
            if h not in hashes_prov:
                prov_col.document(h).set(rec)
                hashes_prov.add(h)
                prov_novos += 1
            else:
                prov_skip += 1

    # ── Posições consolidadas (aba "Carteira") ────────────────────────────
    ws_cart = wb['Carteira']
    pos_col = db.collection(f'users/{USER_UID}/posicoes')
    agora   = datetime.utcnow()
    pos_count = 0
    resumo = {}

    for row in ws_cart.iter_rows(min_row=4, values_only=True):
        ticker = row[0] if len(row) > 0 else None
        if not ticker:
            continue
        ticker = str(ticker).strip().upper()

        # Linha de totais
        if ticker == 'TOTAIS':
            resumo = {
                'custo':        limpar_valor(row[3]),
                'patrimonio':   limpar_valor(row[5]),
                'rentabilidadeRS': limpar_valor(row[6]),
                'rentabilidadePct': limpar_pct(row[7]),
                'proventos':    limpar_valor(row[8]),
                'yieldOnCost':  limpar_pct(row[9]),
                'atualizadoEm': agora,
            }
            db.collection(f'users/{USER_UID}/resumo').document('carteira').set(resumo)
            continue

        # Linha de ativo
        qtd  = row[1]
        if qtd is None or str(qtd).strip() in ('', '0'):
            continue
        try:
            qtd_num = float(str(qtd).replace('.', '').replace(',', '.'))
        except Exception:
            continue

        pos_col.document(ticker).set({
            'ticker':          ticker,
            'classe':          classes.get(ticker, 'Acao'),
            'qtd':             qtd_num,
            'pm':              limpar_valor(row[2]),
            'custo':           limpar_valor(row[3]),
            'cotacao':         limpar_valor(row[4]),
            'patrimonio':      limpar_valor(row[5]),
            'rentabilidadeRS': limpar_valor(row[6]),
            'rentabilidadePct':limpar_pct(row[7]),
            'proventos':       limpar_valor(row[8]),
            'yieldOnCost':     limpar_pct(row[9]),
            'pctAtual':        limpar_pct(row[10]),
            'pctIdeal':        limpar_pct(row[11]),
            'balanceamento':   limpar_valor(row[12]),
            'segmento':        str(row[14]).strip() if row[14] else None,
            'atualizadoEm':    agora,
        })
        pos_count += 1

    print()
    print(f"Lancamentos: {lanc_novos} novos  |  {lanc_skip} ja existiam")
    print(f"Proventos:   {prov_novos} novos  |  {prov_skip} ja existiam")
    print(f"Posicoes:    {pos_count} atualizadas")
    if resumo:
        print(f"Resumo:      Patrimonio R${resumo.get('patrimonio',0):,.2f}  |  Custo R${resumo.get('custo',0):,.2f}")
    print("Sincronizacao concluida!")


if __name__ == '__main__':
    sync()
